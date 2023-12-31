from openpyxl import  load_workbook
from faker import Faker

fake = Faker()

# Read from Excel file 
wb = load_workbook(filename='users_data.xlsx')
ws = wb.active

print(ws['D7'].value)

# Extract all IP addresses
ip_column = ws['F']

for cell in ip_column[1:4]:
    print(cell.value)
    
# Extract by row
for row in ws.values:
    if row[4] == 'Male':
        print(row)

        
# Write to excel file
# To Append values to the active sheet
id = ws.max_row
new_row = [id, fake.first_name(), fake.last_name(), fake.email(), fake.random_element(elements=('Male', 'Female')), fake.ipv4()]
ws.append(new_row)
wb.save(filename='users_data.xlsx')